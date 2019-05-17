"""
problems:
1. Data type include time information
2.
"""

import openpyxl
import copy
import re

# whole new sheet
lst = []
# total columns number
TATAL_COL = 18
# constant zero
ZERO = '0'
# exclude charachter
EXCLUDE = ["", 5]


def isAllZh(s):
    """包含汉字的返回TRUE"""
    for c in s:
        if '\u4e00' <= c <= '\u9fa5':
            return True
    return False


def align(row):
    """
    Align the row into the data that contains TOTAL_COL columns
    """
    # if first col isn't number then put the null string into it
    if row[0] != "" and not is_number(row[0]):
        row.insert(0, "")
    # find the last col(that is the chinese columns) index
    idx = -1
    for i in range(0, len(row)):
        # if row[len(row) - i - 1] != "":
        if row[len(row) - i - 1] not in EXCLUDE:
            idx = len(row) - i - 1
            break
    # insert several null string into the position
    while idx < TATAL_COL:
        row.insert(idx, "")
        idx += 1
    return row


def is_last_number_not_0(row):
    l = list()
    for col in row:
        if col not in EXCLUDE:
            l.append(col)
    # find chinese character
    idx_ch = -1
    for idx in range(0, len(l)):
        if isAllZh(str(l[idx])):
            idx_ch = idx
            break
    # whether the character before the chinese one is a no zero number
    if idx_ch - 1 > 0 and is_number(l[idx_ch - 1]) and str(l[idx_ch - 1]) != ZERO:
        return True
    else:
        return False


def is_number(num):
    pattern = re.compile(r'^[-+]?[-0-9]\d*\.\d*|[-+]?\.?[0-9]\d*$')
    result = pattern.match(str(num))
    if result:
        return True
    else:
        return False


def read_excel(tab):
    # each row of new sheet
    lst2 = list()
    sheet = wb.get_sheet_by_name(tab)

    # for debug only
    idx = 0
    # wash data
    for row in sheet.rows:
        for cell in row:
            if cell.value is not None:
                lst2.append(cell.value)
            else:
                lst2.append("")
        idx += 1
        # print(lst2)
        if is_last_number_not_0(lst2):
            lst.append(align(copy.deepcopy(lst2)))
        lst2.clear()
        # print(lst2)
        # print(lst)
    return lst


def write_excel(path, rows):
    """
    将数据写入excel的tab页中
    """
    wb = openpyxl.Workbook()
    sheet = wb.create_sheet('demo_sheet', index=0)

    # print(isinstance(rows[0], tuple))
    # print(isinstance(rows[0], list))
    # the rows may be a list of tuple/list or a simple list, so need the different deal with it
    if rows is not None and not isinstance(rows[0], tuple) and not isinstance(rows[0], list):
        sheet.append(rows)
    else:
        for row in rows:
            sheet.append(row)

    wb.save(path)


def modify_excel(tab):
    """
    给非0行后面加0
    """
    sheet = wb.get_sheet_by_name(tab)
    x, y = -1, -1

    for i in range(1, sheet.max_row+1):
        for j in range(1, TATAL_COL+2):
            print(sheet.cell(i, j).value)
            if is_number(sheet.cell(i, j).value):
                x = i
                y = j
        if x != -1 and y != -1:
            sheet.cell(x, y+1).value = 0
            x, y = -1, -1

    wb.save("苏浩然法语单词.xlsx")


def diff_sort(tab1, tab2):
    sheet1 = wb.get_sheet_by_name(tab1)
    sheet2 = wb.get_sheet_by_name(tab2)

    s1 = list()
    s2 = list()

    for row in sheet1.rows:
        for cell in row:
            if cell.value is not None and not is_number(cell.value) and not isAllZh(cell.value):
                s1.append(cell.value)
                break
    for row in sheet2.rows:
        for cell in row:
            if cell.value is not None and not is_number(cell.value) and not isAllZh(cell.value):
                s2.append(cell.value)
                break

    s1.sort()
    s2.sort()
    ret_lst = list()
    i = 0
    count = max(len(s1), len(s2))
    while i < count:
        if i < len(s1) and i < len(s2):
            ret_lst.append([s1[i], s2[i]])
        elif i < len(s1):
            ret_lst.append([s1[i], ""])
        else:
            ret_lst.append(["", s2[i]])
        i += 1

    return ret_lst


def diff(tab1, tab2):
    sheet1 = wb.get_sheet_by_name(tab1)
    sheet2 = wb.get_sheet_by_name(tab2)

    s1 = set()
    s2 = set()
    temp_dict = dict()
    ret_lst = list()

    for row in sheet1.rows:
        # print(type(row))
        for cell in row:
            # print(type(cell))
            if cell.value is not None and not is_number(cell.value) and not isAllZh(cell.value):
                s1.add(cell.value)
                temp_dict[cell.value] = row
                break
    for row in sheet2.rows:
        for cell in row:
            if cell.value is not None and not is_number(cell.value) and not isAllZh(cell.value):
                s2.add(cell.value)
                break
    s3 = set()
    s3 = s1 & s2

    temp_lst = list()
    # temp_lst = openpyxl.cell.cell.Cell(sheet1)
    for k in s3:
        temp_lst.append(temp_dict[k])
    # now temp_lst include all rows needed
    # wash data
    for row in temp_lst:
        in_lst = list()
        for cell in row:
            if cell.value is not None:
                in_lst.append(cell.value)
            else:
                in_lst.append("")
        ret_lst.append(copy.deepcopy(in_lst))
        in_lst.clear()
    return ret_lst



wb = openpyxl.load_workbook("e:\\MyCode\\Python\\nalan\\excel\\excel1\\苏浩然法语单词.xlsx")
# 读取非0行并汇总到表格中
read_excel('突击1')
read_excel('突击2')
read_excel('突击3')
read_excel('突击4')
read_excel('突击5')
read_excel('突击6')
read_excel('突击7')
read_excel('突击8')
read_excel('突击9')
# read_excel('突击10')
write_excel("顽固词汇3.xlsx", lst)

# 给非零行加0
# modify_excel("顽固词汇2")

# compare different table and output result
write_excel("相同.xlsx", diff("顽固词汇3", "顽固词汇2"))

# order the compared list and show them
# write_excel("并列排序.xlsx", diff_sort("顽固词汇3", "顽固词汇2"))

